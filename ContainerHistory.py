from selenium import webdriver
from selenium.webdriver.common.keys import Keys  # Allows access to non character keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl
import re

def locate_by_name(web_driver, name):
    """Clicks on something by name."""
    WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.NAME, name))).click()
    # Returns nothing

def get_by_name(web_driver, name):
    """Returns something by name."""
    return WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.NAME, name)))

def locate_by_id(web_driver, id):
    """Clicks on something by id."""
    WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.ID, id))).click()
    # Returns nothing

def get_by_id(web_driver, id):
    """Returns something by id."""
    return WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.ID, id)))

def locate_by_class(web_driver, class_name):
    """Clicks on something by class name."""
    WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, class_name))).click()
    # Returns nothing

def get_by_class(web_driver, class_name):
    """Returns something by class name."""
    return WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, class_name)))

def select_drop(web_driver, id, value):
    """Chooses an option by value in a dropdown box by id."""
    select = Select(web_driver.find_element_by_id(id))
    select.select_by_value(value)
    # Returns nothing

def IT(driver):
    """Create an excel sheet detailing the use of components."""
    inactive_containers = {}  # A dictionary
    def check_container():
        # Get the container number, part number, and location
        temp_container = ((get_by_id(driver, "lblFormTitle").text).split())[2]
        temp_part = get_by_name(driver, "txtPartNo").get_attribute("value")
        temp_location = get_by_id(driver, "pikLocation").text
        # For now, add these to the dictionary
        inactive_containers[temp_container] = [temp_part, temp_location]

        # Check container history
        locate_by_id(driver, "lnkContainerHistory")
        time.sleep(1)
        no_wraps = driver.find_elements_by_class_name("NoWrap")  # This is a list
        # 22 columns, row indices are multiples of 22 (row one starts at index 0)
        row_nums = int(len(no_wraps) / 22)
        # 18 is the first index in the "Last Action" column
        # 14 is the first index in the "Location" column
        act = inact = 0
        for row in range(row_nums):
            if no_wraps[18+(row*22)].text == "Split Container":
                if no_wraps[14+(row*22)].text == "SR RECEIV":
                    continue
                else:
                    act += 1
            elif no_wraps[18+(row*22)].text == "Cycle Complete" or no_wraps[18+(row*22)].text == "Container Move":
                inact += 1
        # Deem the container active or inactive
        if inact >= act:  # Inactive
            if inact == 0:  # A zero ratio means that inact and act are zero
                inactive_containers[temp_container].append(0)
            elif act > 0:
                inactive_containers[temp_container].append(inact/act)
            else:  # If act is zero, but inact is more than zero, the ratio is infinity
                inactive_containers[temp_container].append('inf')
        else:  # Active
            inactive_containers.popitem()  # Remove the most recent addition to the dictionary
        # Go back to the results page
        locate_by_class(driver, "left-arrow-purple")
        locate_by_id(driver, "btnBack_Label")

    # Navigate to inventory tracking menu
    action = ActionChains(driver)
    action.key_down(Keys.CONTROL).send_keys('M').key_up(Keys.CONTROL).perform()
    action = 404
    time.sleep(1)
    action = ActionChains(driver)
    action.send_keys("Inventory Tracking").send_keys(Keys.RETURN).perform()
    action = 404
    time.sleep(1)
    action = ActionChains(driver)
    action.send_keys(Keys.RETURN).perform()

    # Fill out the search criteria
    time.sleep(2)
    input_box = get_by_name(driver, "Layout1$el_385621")
    input_box.clear()
    input_box.send_keys("1/1/2000")
    input_box = get_by_name(driver, "Layout1$el_385622")
    input_box.clear()
    input_box.send_keys("11/25/2019")
    #locs = ["SR03", "SR04", "SR05", "SR06", "SR07", "SR08", "SR09", "SR10", "SR11", "SR12", "SR13", "SR14", "SR15",
    #        "C00", "C01", "C02", "C03", "C04", "C05", "C06", "C07", "C08", "C09", "C10", "C11"]
    # The big list is currently commented out since selenium can occasionally crash for no obvious reason,
    # and these locations have many containers so it would suck if the program crashed in the middle of
    # it all. Should this occur, the program would have to be closed and restarted, which means all progress
    # is restarted.
    locs = ["SR03"]  # For now, use smaller lists with a fraction of the locations
    for loc in locs:  # Cycle through locations
        # Find the location box and enter in the name
        input_box = get_by_id(driver, "Layout1_el_6102")
        input_box.clear()
        input_box.send_keys(loc)
        # Find and click the search button
        locate_by_id(driver, "Layout1_el_56")
        time.sleep(1)

        # Find length of results
        links = driver.find_elements_by_xpath("//a[@href]")  # Get every link
        loc_links = 0  # This variable stores the amount of containers in one location
        for link in links:  # Find all instances of a container link
            if re.search("ContainerForm", link.get_attribute("href")):
                loc_links += 1
        if loc_links == 0:  # If no results come up
            continue  # Go to the next location

        for index in range(loc_links):  # Cycle through containers in one location
            links = driver.find_elements_by_xpath("//a[@href]")  # "Refresh" the links
            encountered = 0
            for link in links:  # Cycle through all container links in one location
                if re.search("ContainerForm", link.get_attribute("href")):
                    if encountered == index:
                        link.click()  # Enter container
                        check_container()
                        break
                    encountered += 1

    # Write data to new excel sheet
    wb_obj = openpyxl.Workbook()
    sheet_obj = wb_obj.active
    headers = ["Container Number", "Part Number", "Location", "Activity Ratio (inactive/active)"]
    for i in range(1, 5):  # Write in headers
        sheet_obj.cell(row=1, column=i).value = headers[i-1]
    for index, key in enumerate(inactive_containers):
        sheet_obj.cell(row=index+2, column=1).value = key
        sheet_obj.cell(row=index+2, column=2).value = inactive_containers[key][0]
        sheet_obj.cell(row=index+2, column=3).value = inactive_containers[key][1]
        sheet_obj.cell(row=index+2, column=4).value = inactive_containers[key][2]
    # This creates a new excel workbook in the program's directory
    wb_obj.save("Inactive Containers List.xlsx")

try:
    # Getting into Plex
    driver = webdriver.Chrome("chromedriver.exe")
    driver.get("https://www.plexonline.com/modules/systemadministration/login/index.aspx?")
    driver.find_element_by_name("txtUserID").send_keys("w.Andre.Le")
    driver.find_element_by_name("txtPassword").send_keys("OokyOoki2")
    driver.find_element_by_name("txtCompanyCode").send_keys("wanco")
    locate_by_id(driver, "btnLogin")
    # Since logging in through selenium opens up a new window instead of changing the current login screen,
    # go to the new screen
    driver.switch_to.window(driver.window_handles[1])
    time.sleep(3)
    IT(driver)
except Exception as e:
    print("An error was encountered:")
    print(e)
finally:  # Whether the operation was erroneous or successful, quit the driver
    driver.quit()
